from datetime import datetime
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os.path
from os import path
import collections
from openpyxl.styles import Font,Alignment
from openpyxl.utils import get_column_letter
# Reading an excel file using Python
#import xlrd
#
## Give the location of the file
#loc = ("Example3.xlsx")
#
## To open Workbook
#wb = xlrd.open_workbook(loc)
#sheet = wb.sheet_by_index(0)
#
## For row 0 and column 0
#a=sheet.cell_value("E8")


"""
    File WriteXls.py have the role to create XLSX and add in it
"""


class XlsWriter:
    """ The main class that will Work with XLSX.
     You can add any nationality or Brand in the next 2 dicts."""

    Nationality = {
        "eu-fr": "",
        "eu-sk": "",
        "eu-de": "",
        "eu-gb": "",
        "eu-ro": "",
        "eu-it": "",
        "Switzerland": "",
        "Other":""
    }

    Aux={
        "All Nat": "",
        "All Brand":""
    }
    Vehicles={
        "Cars":[12,5],
        "Trucks":[13,5],
        "All Vehicles":[14,5]
    }

    Hours={
    }

    Brand = {
        "citroen": "",
        "peugeot":"",
        "volvo": "",
        "ford": "",
        "tesla":"",
        "Alfa Romeo": "",
        "renault": "",
        "skoda": "",
        "VW": "",
        "Other":""
    }


    def __init__(self,PathToSaveXlsx):
        """Initialize the variables"""
        self.PathToSaveXlsx=PathToSaveXlsx
        self.set_place()


    def create_new_xls(self,date):
        """Create the XLSX with all the formula and table"""
        try:
            time = int(date) / 1000

            dateget = datetime.utcfromtimestamp(time).strftime('%d_%m_%Y')
            FileName=self.PathToSaveXlsx+str(dateget)+"_OpenAlpr.xlsx"
            if path.exists(FileName):
                pass
            else:
                self.workbook = Workbook()

                self.sheet = self.workbook.active

                self.sheet.column_dimensions['E'].width=15

                self.sheet.merge_cells('F7:AC7')
                self.sheet.cell(7,6,"Time").font = Font(bold=True)
                self.sheet.cell(7, 6, "Time").alignment = Alignment(horizontal='center')
                #aa.value="Time"

                self.sheet.cell(8,5, 'Day').font = Font(bold=True)

                #write date
                #import time
                #DateNow = time.strftime("%d-%m-%Y")
                #from datetime import date
                #import calendar
                #my_date = date.today()
                #day=calendar.day_name[my_date.weekday()]
                self.sheet.cell(9,5,dateget)

                #write hours
                for hours in self.Hours:
                    #atentie- pentru fiecare coloana cu ora, creez si suma mai jos pt fiecare cars si truck
                    self.sheet.cell(self.Hours[hours][0],self.Hours[hours][1],hours)

                    #Creez sumele pe ore
                    #vehicle
                    column_letter1=get_column_letter(self.Hours[hours][1])
                    SumVehicle="("+column_letter1+str(self.Vehicles['Cars'][0])+":"+column_letter1+str(self.Vehicles['Trucks'][0])+")"
                    self.sheet.cell(self.Vehicles['All Vehicles'][0],self.Hours[hours][1],'=SUM'+SumVehicle)

                    #nationality

                    SumNat="("
                    for nat in self.Nationality:
                        column_letter1 = get_column_letter(self.Hours[hours][1])
                        SumNat+=column_letter1+str(self.Nationality[nat][0])+":"
                    SumNat = SumNat[:-1]
                    SumNat = SumNat + ")"
                    self.sheet.cell(self.Aux['All Nat'][0], self.Hours[hours][1], '=SUM' + SumNat)

                    #Brand
                    SumBrand = "("
                    for nat in self.Brand:
                        column_letter1 = get_column_letter(self.Hours[hours][1])
                        SumBrand += column_letter1 + str(self.Brand[nat][0]) + ":"
                    SumBrand = SumBrand[:-1]
                    SumBrand = SumBrand + ")"
                    self.sheet.cell(self.Aux['All Brand'][0], self.Hours[hours][1], '=SUM' + SumBrand)




                self.sheet.cell(self.Hours['23:00'][0],(self.Hours['23:00'][1])+1,"TOTAL").font = Font(bold=True)


                #write vehicles
                self.sheet.cell(11,5, 'Vehicles').font = Font(bold=True)

                #Vehicle
                for vehicle in self.Vehicles:
                    self.sheet.cell(self.Vehicles[vehicle][0],self.Vehicles[vehicle][1], vehicle)
                    SumVehicleByType="("
                    #creez suma de total din partea dreapta pt fiecare linie mare
                    for hours in self.Hours:
                        column_letter1 = get_column_letter(self.Hours[hours][1])
                        SumVehicleByType +=  column_letter1 + str(self.Vehicles[vehicle][0]) + ":"
                    SumVehicleByType = SumVehicleByType[:-1]
                    SumVehicleByType=SumVehicleByType+")"
                    self.sheet.cell(self.Vehicles[vehicle][0],(self.Hours['23:00'][1])+1,'=SUM'+SumVehicleByType).font = Font(bold=True)

                #Nationality
                self.sheet.cell(self.NationalityPlace[0],self.NationalityPlace[1], 'Nationality').font = Font(bold=True)
                for nationality in self.Nationality:
                    self.sheet.cell(self.Nationality[nationality][0],self.Nationality[nationality][1], nationality)
                    SumNationality = "("
                    # creez suma de total din partea dreapta pt fiecare linie mare
                    for hours in self.Hours:
                        column_letter1 = get_column_letter(self.Hours[hours][1])
                        SumNationality += column_letter1 + str(self.Nationality[nationality][0]) + ":"
                    SumNationality = SumNationality[:-1]
                    SumNationality = SumNationality + ")"
                    self.sheet.cell(self.Nationality[nationality][0], (self.Hours['23:00'][1]) + 1, '=SUM' + SumNationality).font = Font(bold=True)

                self.sheet.cell(self.Aux["All Nat"][0],self.Aux["All Nat"][1],"All Nat")

                #linia de nationality cu suma totala din dreapta
                SumNationality = "("
                # creez suma de total din partea dreapta pt fiecare linie mare
                for hours in self.Hours:
                    column_letter1 = get_column_letter(self.Hours[hours][1])
                    SumNationality += column_letter1 + str(self.Aux["All Nat"][0]) + ":"
                SumNationality = SumNationality[:-1]
                SumNationality = SumNationality + ")"
                self.sheet.cell(self.Aux["All Nat"][0], (self.Hours['23:00'][1]) + 1, '=SUM' + SumNationality).font = Font(bold=True)

                #Brand
                self.sheet.cell(self.BrandPlace[0],self.BrandPlace[1], 'Brand').font = Font(bold=True)
                for brand in self.Brand:
                    self.sheet.cell(self.Brand[brand][0],self.Brand[brand][1], brand)
                    SumBrand = "("
                    # creez suma de total din partea dreapta pt fiecare linie mare
                    for hours in self.Hours:
                        column_letter1 = get_column_letter(self.Hours[hours][1])
                        SumBrand += column_letter1 + str(self.Brand[brand][0]) + ":"
                    SumBrand = SumBrand[:-1]
                    SumBrand = SumBrand + ")"
                    self.sheet.cell(self.Brand[brand][0], (self.Hours['23:00'][1]) + 1, '=SUM' + SumBrand).font = Font(bold=True)
                self.sheet.cell(self.Aux["All Brand"][0], self.Aux["All Brand"][1], "All Brand")
                SumBrand = "("
                # creez suma de total din partea dreapta pt fiecare linie mare
                for hours in self.Hours:
                    column_letter1 = get_column_letter(self.Hours[hours][1])
                    SumBrand += column_letter1 + str(self.Aux["All Brand"][0]) + ":"
                SumBrand = SumBrand[:-1]
                SumBrand = SumBrand + ")"
                self.sheet.cell(self.Aux["All Brand"][0], (self.Hours['23:00'][1]) + 1, '=SUM' + SumBrand).font = Font(bold=True)


                self.workbook.save(filename=FileName)
            return FileName
        except Exception as e:
            raise e



    def set_place(self):
        """Initialize the hours and all from dictionaries."""
        #hours
        try:
            for index in range(0,24):
                if int(index)<10:
                    self.Hours["0"+str(index) + ":00"] = [8, index + 6]  # 8 si +6 sunt pozitiile in csv
                else:
                    self.Hours[str(index)+":00"]=[8,index+6]#8 si +6 sunt pozitiile in csv


            raw_start=16
            self.NationalityPlace=[raw_start,5]
            for nat in self.Nationality:
                raw_start=raw_start+1
                self.Nationality[nat]=[raw_start,5]
            raw_start+=1
            self.Aux["All Nat"]=[raw_start,5]

            raw_start=raw_start+2
            self.BrandPlace=[raw_start,5]
            for brand in self.Brand:
                raw_start=raw_start+1
                self.Brand[brand]=[raw_start,5]
            raw_start += 1
            self.Aux["All Brand"] = [raw_start, 5]
        except Exception as e:
            raise e


    def set_value_increment(self,jsonget,type):
        """Increment values depending on type detected and hour"""
        from openpyxl import load_workbook
        try:
            #self.workbook.sheetnames['Sheet']
            #check if the file for this date exist or not -> create

            NameFile=self.create_new_xls(jsonget['rest']['date'])
            self.workbook = load_workbook(filename=NameFile)
            self.sheet = self.workbook.active
            #get timestamp
            time=int(jsonget['rest']['date'])/1000
            dateHour=datetime.utcfromtimestamp(time).strftime('%H')
            dateHour=dateHour+":00"
            getSpaceColumn=self.Hours[dateHour][1]

            #vehicle type
            getSpaceRow=self.Vehicles[(type+"s").title()][0]

            value=self.sheet.cell(getSpaceRow,getSpaceColumn).value
            if value==None:
                value=1
            else:
                value=value+1
            self.sheet.cell(getSpaceRow,getSpaceColumn).value=value

            #nationality
            nationality=jsonget['rest']['rest']['best_region']
            if nationality in self.Nationality:
                getSpaceRow = self.Nationality[(nationality)][0]
            else:
                getSpaceRow=self.Nationality['Other'][0]

            value = self.sheet.cell(getSpaceRow, getSpaceColumn).value
            if value == None:
                value = 1
            else:
                value = value + 1
            self.sheet.cell(getSpaceRow, getSpaceColumn).value = value
            #brand
            brand=jsonget['rest']['rest']['vehicle']['make'][0]['name']
            if brand in self.Brand:
                getSpaceRow = self.Brand[(brand)][0]
            else:
                getSpaceRow = self.Brand['Other'][0]

            value = self.sheet.cell(getSpaceRow, getSpaceColumn).value
            if value == None:
                value = 1
            else:
                value = value + 1
            self.sheet.cell(getSpaceRow, getSpaceColumn).value = value

            self.workbook.save(NameFile)
        except Exception as e:
            raise e

##NEED TO DO BY TYME TODO
#every insert will be by the timestamp of the json.
#check the json
#a=XlsWriter()